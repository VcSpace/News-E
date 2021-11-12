import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import json
import time

headers = {
    'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36',
}

class FOX(object):
    def __init__(self):
        pass

    def request(self):
        self.bs_url = 'https://www.foxbusiness.com/markets'
        for _ in range(10):
            try:
                self.bs_data = requests.get(self.bs_url, headers=headers)
                self.bs_soup = BeautifulSoup(self.bs_data.text, "lxml")
                break
            except:
                continue

    def request2(self):
        self.en_url = 'https://www.foxbusiness.com/economy'
        for _ in range(10):
            try:
                self.en_data = requests.get(self.en_url, headers=headers)
                self.en_soup = BeautifulSoup(self.en_data.text, "lxml")
                break
            except:
                continue

    def have_video(self, href):
        index = href.find('video')
        if index == -1:
            href = "https://www.foxbusiness.com" + href
        return href

    def get_bs_news(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.create_sheet("Fox")

        t_row = 1
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="东方财富")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        topdata = self.bs_soup.select('#wrapper > div > div.page-content > main > div.collection.collection-big-top > article > div.info > header > h2')
        for news in topdata:
            m_href = news.find('a')['href']
            m_href = self.have_video(m_href)
            m_title = news.find('a').get_text()
            sheet.cell(row=t_row, column=t_col, value=m_title)
            sheet.cell(row=t_row, column=t_col + 1, value=m_href)
            t_row = t_row + 1

        top2data = self.bs_soup.select('#wrapper > div > div.page-content > main > section.collection.collection-2-across > div')
        for data in top2data:
            news = data.find_all(class_='title')
            for m_news in news:
                m_news = m_news.find('a')
                m_href = m_news['href']
                m_href = self.have_video(m_href)
                m_title= m_news.get_text()
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                t_row = t_row + 1

        datalist = self.bs_soup.select('#wrapper > div > div.page-content > main > div.collection.collection-river.content > div')
        i = 1
        for data in datalist:
            news = data.find_all('a')
            for m_news in news:
                if i % 2 == 0:
                    m_href = m_news['href']
                    m_title = m_news.get_text()
                    m_href = self.have_video(m_href)
                    sheet.cell(row=t_row, column=t_col, value=m_title)
                    sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                    t_row = t_row + 1
                i = i + 1

        try:
            wb.save(self.xlsxname)
        except Exception:
            print("FOX Save Error = 1")

    def main(self, file_name):
        self.xlsxname = file_name
        Fox.request()
        Fox.get_bs_news()

Fox = FOX()
