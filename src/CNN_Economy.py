import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import json
import time

headers = {
    'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36',
}

class CNN(object):
    def __init__(self):
        pass

    def request(self):
        self.url = 'https://edition.cnn.com/business/'
        for _ in range(10):
            try:
                self.data = requests.get(self.url, headers=headers)
                self.soup = BeautifulSoup(self.data.text, "lxml")
                break
            except:
                continue

    def request2(self):
        #self.en_url = 'https://edition.cnn.com/business/economy'
        self.en_url = 'https://edition.cnn.com/data/ocs/container/coverageContainer_7FCD2EF6-BE29-BE10-44E3-34EB048CF28E:list-xs/views/containers/common/container-manager.html'
        for _ in range(10):
            try:
                self.en_data = requests.get(self.en_url, headers=headers)
                self.en_soup = BeautifulSoup(self.en_data.text, "lxml")
                break
            except:
                continue

    def get_business(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name('Cnn')
        t_row = 1
        t_col = 1

        sheet.cell(row=t_row + 0, column=t_col + 0, value="CNN商业")
        sheet.cell(row=t_row + 1, column=t_col + 0, value="新闻标题")
        sheet.cell(row=t_row + 1, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row + 1, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row + 1, column=t_col + 3, value="新闻时间")
        t_row = t_row + 2

        datalist = self.soup.select('#us-zone-1 > div.l-container > div > div.column.zn__column--idx-0 > ul')
        for data in datalist:
            news = data.find_all('a')
            for m_new in news:
                m_href = "https://edition.cnn.com" + m_new['href']
                m_title = m_new.get_text()
                if m_title == '':
                    continue
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                t_row = t_row + 1

        datalist2 = self.soup.select('#us-zone-1 > div.l-container > div > div.column.zn__column--idx-1 > ul')
        for data in datalist2:
            news = data.find_all('a')
            for m_new in news:
                m_href = "https://edition.cnn.com" + m_new['href']
                m_title = m_new.get_text()
                if m_title == '':
                    continue
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                t_row = t_row + 1

        try:
            wb.save(self.xlsxname)
        except Exception:
            print("CNN Save Error = 1")


    def create_file(self, file_name):
        self.xlsxname = file_name
        wb = Workbook()
        ws = wb['Sheet']
        wb.remove(ws)
        sheet = wb.create_sheet("Cnn")

        try:
            wb.save(self.xlsxname)
        except Exception:
            print("CNN create_error = 1")

    def get_economy(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name("Cnn")
        t_row = sheet.max_row + 2
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="CNN经济")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        datalist2 = self.en_soup.select('#coverageContainer_7FCD2EF6-BE29-BE10-44E3-34EB048CF28E > ul')
        for data in datalist2:
            news = data.find_all('a')
            for m_new in news:
                m_href = "https://edition.cnn.com" + m_new['href']
                m_title = m_new.get_text()
                if m_title == '':
                    continue
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                t_row = t_row + 1

        try:
            wb.save(self.xlsxname)
        except Exception:
            print("CNN Save Error = 2")

    def main(self, file_name):
        Cnn.create_file(file_name)
        Cnn.request()
        Cnn.get_business()
        Cnn.request2()
        Cnn.get_economy()


Cnn = CNN()
